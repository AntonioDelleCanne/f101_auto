import pandas as pd
import os
import sys
from docx import Document
from pathlib import Path
from nbdev.showdoc import doc
import ipywidgets as widgets
from ipywidgets import interact, fixed, FileUpload
import numpy as np
import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook
from IPython.display import display
from ipywidgets import HTML
from functools import partial
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
import shutil

from database import *

# TODO fix formulas: overall satisfaction should go from B_scorerow:lastcolletter_scorerows
# TODO fixe dived by 0 error in formulas
# TODO put document links
# TODO create script that retrieves document links
class XLSXDoc:
    
    def __init__(self, db, xlsx_path=xlsx_path, month_year="JUN2021", ws_name='HR3-4 ', template_path=template_path, t_ws_name='HR3-4 '):
        self.xlsx_path = xlsx_path
        self.month_year = month_year
        self.file_name = f'NET UK - QMS - {month_year} - Indicators HR3-HR4.xlsx'
        self.file_path = self.xlsx_path/self.file_name
        self.xlsx_path.mkdir(parents=True, exist_ok=True)
        self.ws_name = ws_name
        self.template_path = template_path
        self.t_ws_name = t_ws_name
        self.db = db
        db.listen(self)
        self.regenerate()
    
    def regenerate(self):
        '''
        Regenerate the excel according to the database
        '''
        wb = Workbook()
        ws = wb.active
        ws.title = self.ws_name
        
        db_dict = self.db.get_db()
        scores = db_dict['scores']
        n_employees = db_dict['n_employees']
        
        xlsx_structure = {
            'A1': 'Provided courses 2020-2021',
            'A2': 'for each course is reported the average of the opinions expressed with rating from 0 to 5'
        }
        
        for k, v in xlsx_structure.items():
            ws[k] = v

        # add the scores
        for i,r in enumerate(dataframe_to_rows(scores, index=True, header=True)):
            if(i != 1):
                ws.append(r)
        ws['A3'] = 'HR4'

        # popualate the score calcuation (with eqatuions)
        score_table_end = 2 + len(list(dataframe_to_rows(scores, index=True, header=True)))
        ws.cell(score_table_end,1, value = 'Course average rating')
        ws = self.__regenerate_formulas(ws)

        # add the docx links, a list of links for each course
        courses_docx = self.db.get_courses_docx()
        ws.cell(score_table_end + 1, 1, value='General Notes on the effectiveness of the course')
        for i, c in enumerate(scores.columns):
            if c in courses_docx.keys():
                urls = courses_docx[c]
                cell_value = ''
                for url in urls:
                    fname = url.stem
                    path = 'docx'
                    fpath = path + '/' + fname + '.docx'
                    cell_value = cell_value + str(fpath) + ' ; '
                cell_value = cell_value[:-3]
                ws.cell(score_table_end + 1, 2+i, value=cell_value)
            

        hr3_row = score_table_end + 3
        ws.cell(hr3_row, 1, value='HR3')

        # add 

        course_nr_row = hr3_row+1
        ws.cell(course_nr_row, 1, value='provided courses (nr.)')
        ws.cell(course_nr_row, 2, value='=20-COUNTIFS(3:3, "Column*")')

        n_people_row = hr3_row+2
        ws.cell(n_people_row, 1, value='Number of people affected')
        ws.cell(n_people_row, 2, value='=COUNTA(Tabella2[HR4])')

        n_employees_row = hr3_row+3

        ws.cell(n_employees_row, 1, value='Number of employees')
        ws.cell(n_employees_row, 2, value=n_employees)
        ws.cell(n_employees_row, 3, value='* taken as the current total')


        summary_row = hr3_row + 5
        year = 2020
        ws.cell(summary_row, 1, value=f'Summary {year}')
        ws.cell(summary_row, 2, value='Indexes')
        ws.cell(summary_row, 3, value='Threshold')

        ws.cell(summary_row+1, 1, value='HR3 - coverage of staff')
        ws.cell(summary_row+1, 2, value='=B15/B16')

        ws.cell(summary_row+2, 1, value='HR4 - overall satisfaction')
        ws.cell(summary_row+2, 2, value='=AVERAGE(B10:R10)')
        
        self.save_changes_to_file(wb)
        self.__aplly_template_style()
        
        self.regenerate_download_direcotry()
        
        return wb, ws
    
    
    def regenerate_download_direcotry(self):
        # TODO put data in download_temp
        docx_paths = self.db.get_courses_docx()
        for courses in docx_paths.values():
            for src in courses:
                dest = download_temp_docx/(str(src.stem) + '.docx')
                dest.touch()
                shutil.copy(src, dest)
        xlsx_dest = download_temp/(str(self.file_path.stem) + '.xlsx')
        shutil.copy(self.file_path, xlsx_dest)
        shutil.make_archive(download_dir, 'zip', download_temp)
        
    
    def get_wb_ws(self):
        '''
        returns a the workbook corresponding to the actual file
        and the worksheet view of it
        '''
        wb = load_workbook(self.file_path)
        ws = wb[self.ws_name]
        return wb, ws
        
    def get_courses(self):
        wb, ws = self.get_wb_ws()
        df = pd.DataFrame(ws.values).set_index(0)
        courses = df.iloc[2,:].dropna()
        return list(courses)
    
    def user_exists(self, name):
        wb, ws = self.get_wb_ws()
        df = pd.DataFrame(ws.values).set_index(0)
        row_n = df.index.get_loc('Course average rating') + 1
        if(name in list(df.index)[3:row_n]):
            return True
        return False
    
    def course_exists(self, name):
        if(name in self.get_courses()):
            return True
        return False
    
    def add_user(self, name):
        wb, ws = self.get_wb_ws()
        if self.user_exists(name):
            raise ValueError('user already exists')
        ws.insert_rows(4)
        ws['A4'] = name
        self.__regenerate_formulas(ws)
        self.save_changes_to_file(wb)
        return ws
    
    def add_score(self, user, course, score):
        if course not in self.get_courses():
            raise ValueError('Course does not exist')
        if not self.user_exists(user):
            self.add_user(user)
        wb, ws = self.get_wb_ws()
        df = pd.DataFrame(ws.values).set_index(0)
        row_n = df.index.get_loc(user) + 1
        col_n = pd.Index(df.iloc[2]).get_loc(course) + 2
        ws.cell(row_n,col_n, value = score)
        self.save_changes_to_file(wb)
      
    def add_course(self, course):
        if self.course_exists(course):
            raise ValueError('Course already exists')
        wb, ws = self.get_wb_ws()
        df = pd.DataFrame(ws.values).set_index(0)
        
        row_n = 3
        col_n = len(pd.Index(df.iloc[2])) + 2
        
        ws.cell(row_n,col_n, value = course)
        self.__regenerate_formulas(ws)
        self.save_changes_to_file(wb)
    
    def show_df(self):
        xlsx_file = self.file_path
        wb = load_workbook(xlsx_file)
        ws = wb['HR3-4 ']
        display(pd.DataFrame(ws.values).set_index(0))
            
    def save_changes_to_file(self, wb):
#         wb.save(self.file_path)
        openpyxl.writer.excel.save_workbook(wb, self.file_path)
        
    def notify(self):
        self.regenerate()
    
    def __regenerate_formulas(self, ws):
        df = pd.DataFrame(ws.values).set_index(0)
        row_n = df.index.get_loc('Course average rating') + 1
        courses = df.iloc[2,:].dropna()
        for i in range(len(courses)):
            i0 = 2
            cell = ws.cell(row_n,i0+i)
            interval = f'{cell.column_letter}4:{cell.column_letter}{row_n-1}'
            f_new = f'=IF(ISERROR(AVERAGE({interval})) = TRUE, "", AVERAGE({interval}))'
            ws.cell(row_n,i0+i, value=f_new)
        return ws
    
    def __aplly_template_style(self):
        
        db_dict = self.db.get_db()
        scores = db_dict['scores']
        n_employees = db_dict['n_employees']
        score_table_end = 2 + len(list(dataframe_to_rows(scores, index=True, header=True)))
        
        t_wb = load_workbook(self.template_path)
        t_ws = t_wb[self.t_ws_name]
        wb, ws = self.get_wb_ws()
        # contais the styles with the corresponding cell position in the template sheet
        styles = {'provided_courseA1': (1,1),
                  'descriptionB1': (2,1),
                  'HR4_head': (3,1),
                  'HR4_rows': (4,1),
                  'HR4_body': (4,2),
                  'HR4_avg_head': (10,1),
                  'HR4_avg_scores': (10,2),
                  'general_notes': (11,1),
                  'general_notes_cell': (11,2),
                  'HR3_head': (13,1),
                  'HR3_rows': (14,1),
                  'HR3_body': (14,2),
                  'HR3_nempl': (16,2),
                  'HR3_asterisk_note': (16,3),
                  'summary_head': (18,1),
                  'summary_rows': (19,1),
                  'summary_body': (19,2),
                  'summary_thr': (19,3)
                 }
        
        rdims = {}
        for i in range(20):
            rdims[i] = None
            
        for k, rd in t_ws.row_dimensions.items():
            rdims[k] = rd.height
#             i = int(k) - 3
#             if i >=1:
#                 ws.row_dimensions[i].height = rd.height
        
        # apply the styles to the cells of the regenerated sheet
        ln=1
        s = 'provided_courseA1'
        ws.row_dimensions[ln].height = rdims[styles[s][0]]
        self.__apply_cell_style(t_ws.cell(*styles[s]), ws.cell(1,1))
        
        ln=2
        s = 'descriptionB1'
        ws.row_dimensions[ln].height = rdims[styles[s][0]]
        self.__apply_cell_style(t_ws.cell(*styles[s]), ws.cell(2,1))
        
        
        for i in range(1,len(scores.columns) + 2):
            ln = i
            s = 'HR4_head'
            ws.row_dimensions[ln].height = rdims[styles[s][0]]
            self.__apply_cell_style(t_ws.cell(*styles[s]), ws.cell(3,i))
        
        for i in range(4, score_table_end):
            ln = i
            s = 'HR4_rows'
            ws.row_dimensions[ln].height = rdims[styles[s][0]]
            self.__apply_cell_style(t_ws.cell(*styles[s]), ws.cell(i,1))
            for j in range(2,len(scores.columns) + 2):
                self.__apply_cell_style(t_ws.cell(*styles['HR4_body']), ws.cell(i,j))
        
        ln = score_table_end
        s = 'HR4_avg_head'
        ws.row_dimensions[ln].height = rdims[styles[s][0]]
        self.__apply_cell_style(t_ws.cell(*styles[s]), ws.cell(score_table_end,1))
        for i in range(2,len(scores.columns) + 2):
            self.__apply_cell_style(t_ws.cell(*styles['HR4_avg_scores']), ws.cell(score_table_end,i))
        
        ln = score_table_end + 1
        s = 'general_notes'
        ws.row_dimensions[ln].height = rdims[styles[s][0]]
        self.__apply_cell_style(t_ws.cell(*styles[s]), ws.cell(score_table_end+1,1))
        for i in range(2,len(scores.columns) + 2):
            self.__apply_cell_style(t_ws.cell(*styles['general_notes_cell']), ws.cell(score_table_end+1,i))
        
        # TODO if needed chsnge row dimensions for folowing tables too with
#         ln = score_table_end + 1
#         s = 'general_notes'
#         ws.row_dimensions[ln].height = rdims[styles[s][0]]
        
        hr3_head = score_table_end+3
        for i in range(1, 3):
            self.__apply_cell_style(t_ws.cell(*styles['HR3_head']), ws.cell(hr3_head,i))
        
        for i in range(hr3_head+1, hr3_head+4):
            self.__apply_cell_style(t_ws.cell(*styles['HR3_rows']), ws.cell(i,1))
            if(i < hr3_head+3):
                self.__apply_cell_style(t_ws.cell(*styles['HR3_body']), ws.cell(i,2))
            else:
                self.__apply_cell_style(t_ws.cell(*styles['HR3_nempl']), ws.cell(i,2))
                self.__apply_cell_style(t_ws.cell(*styles['HR3_asterisk_note']), ws.cell(i,3))
        
        summary_head = hr3_head + 5
        
        for i in range(1,4):
            self.__apply_cell_style(t_ws.cell(*styles['summary_head']), ws.cell(summary_head,i))
        
        for i in range(summary_head+1, summary_head+3):
            self.__apply_cell_style(t_ws.cell(*styles['summary_rows']), ws.cell(i,1))
            self.__apply_cell_style(t_ws.cell(*styles['summary_body']), ws.cell(i,2))
            self.__apply_cell_style(t_ws.cell(*styles['summary_thr']), ws.cell(i,3))
        
        for k, cd in t_ws.column_dimensions.items():
            ws.column_dimensions[k].width = cd.width
        
        # TODO change dynamically based on courses
#         for k, rd in t_ws.row_dimensions.items():
#             i = int(k) - 3
#             if i >=1:
#                 ws.row_dimensions[i].height = rd.height
        
        self.save_changes_to_file(wb)
    
    def __apply_cell_style(self, cell_src, cell_dest):
        cell_dest.font = copy(cell_src.font)
        cell_dest.fill = copy(cell_src.fill)
        cell_dest.border = copy(cell_src.border)
        cell_dest.alignment = copy(cell_src.alignment)
        cell_dest.number_format = copy(cell_src.number_format)
        cell_dest.protection = copy(cell_src.protection)
        return cell_dest
    